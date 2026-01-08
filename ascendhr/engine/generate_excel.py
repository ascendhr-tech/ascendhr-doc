#!/usr/bin/env python3
"""Generate Player Card Calculation Logic Excel file for dev team."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()

# ===== SHEET 1: README =====
ws1 = wb.active
ws1.title = "README"

# Styles
header_font = Font(bold=True, size=14, color="FFFFFF")
header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
subheader_font = Font(bold=True, size=11)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# README content
readme_data = [
    ["Player Card Calculation Logic - Summary"],
    [""],
    ["This Excel file summarizes the calculation formulas for the Player Card system."],
    ["Use this as a reference for implementing the calculation engine."],
    [""],
    ["SHEETS:"],
    ["1. README - This overview"],
    ["2. Attribute_Library - All attributes with weights"],
    ["3. Position_Requirements - Sample position config"],
    ["4. Overall_Score - Calculator with formulas"],
    ["5. Fit_Score - Position fit calculator"],
    ["6. Example - Complete worked example"],
    [""],
    ["KEY CONCEPTS:"],
    ["• Attribute Weight: Low (×1.0), Medium (×1.25), High (×1.5) - affects Overall Score"],
    ["• Position Priority: ★ (×1), ★★ (×2), ★★★ (×3) - affects Fit Score"],
    ["• Scale: 1-5 Low, 6-10 Developing, 11-15 Proficient, 16-20 Exceptional"],
    ["• Overall Score: Weighted average × 5, range 0-100"],
    ["• Fit Score: Priority-weighted average of attribute scores, range 0-100%"],
    [""],
    ["Version: 2.0"],
    ["Date: 2024-12-25"],
]

for row_num, row_data in enumerate(readme_data, 1):
    ws1.cell(row=row_num, column=1, value=row_data[0])
    if row_num == 1:
        ws1.cell(row=row_num, column=1).font = Font(bold=True, size=16)
    elif ":" in str(row_data[0]) and row_data[0].endswith(":"):
        ws1.cell(row=row_num, column=1).font = subheader_font

ws1.column_dimensions['A'].width = 70

# ===== SHEET 2: Attribute Library =====
ws2 = wb.create_sheet("Attribute_Library")

attr_headers = ["Attribute", "Category", "Weight", "Multiplier", "Description"]
for col, header in enumerate(attr_headers, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

attributes = [
    ["Leadership", "Core", "High", 1.5, "Guide and inspire teams"],
    ["Teamwork", "Core", "High", 1.5, "Collaborative effectiveness"],
    ["Communication", "Core", "Medium", 1.25, "Clear expression"],
    ["Adaptability", "Core", "Medium", 1.25, "Flexibility in change"],
    ["Determination", "Core", "Medium", 1.25, "Persistence and drive"],
    ["Coding", "Technical", "High", 1.5, "Programming proficiency"],
    ["System Design", "Technical", "High", 1.5, "Architecture skills"],
    ["Problem Solving", "Technical", "Medium", 1.25, "Analytical thinking"],
    ["Visual Design", "Technical", "High", 1.5, "Aesthetic sense"],
    ["UX Thinking", "Technical", "Medium", 1.25, "User-centered design"],
    ["Data Analysis", "Technical", "Medium", 1.25, "Data interpretation"],
    ["Negotiation", "Business", "High", 1.5, "Deal-making skills"],
    ["Closing", "Business", "High", 1.5, "Sales conversion"],
    ["Client Relations", "Business", "Medium", 1.25, "Relationship building"],
    ["Strategic Thinking", "Business", "Medium", 1.25, "Long-term planning"],
    ["Creativity", "Creative", "Low", 1.0, "Innovation and ideas"],
    ["Prototyping", "Creative", "Low", 1.0, "Rapid iteration"],
    ["Design Thinking", "Creative", "Medium", 1.25, "Human-centered approach"],
]

for row_num, attr in enumerate(attributes, 2):
    for col, val in enumerate(attr, 1):
        cell = ws2.cell(row=row_num, column=col, value=val)
        cell.border = thin_border

for col in range(1, 6):
    ws2.column_dimensions[get_column_letter(col)].width = [20, 12, 10, 12, 30][col-1]

# ===== SHEET 3: Position Requirements =====
ws3 = wb.create_sheet("Position_Requirements")

pos_headers = ["Position", "Attribute", "Priority", "Priority Weight", "Min Value"]
for col, header in enumerate(pos_headers, 1):
    cell = ws3.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

pos_reqs = [
    ["Sr. Backend Developer", "Coding", "Critical ★★★", 3, 15],
    ["Sr. Backend Developer", "System Design", "Critical ★★★", 3, 13],
    ["Sr. Backend Developer", "Teamwork", "Critical ★★★", 3, 14],
    ["Sr. Backend Developer", "Leadership", "Important ★★", 2, 12],
    ["Sr. Backend Developer", "Problem Solving", "Important ★★", 2, 12],
    ["Sr. Backend Developer", "Communication", "Nice-to-have ★", 1, 10],
    ["", "", "", "", ""],
    ["Product Manager", "Leadership", "Critical ★★★", 3, 15],
    ["Product Manager", "Communication", "Critical ★★★", 3, 16],
    ["Product Manager", "Strategic Thinking", "Critical ★★★", 3, 14],
    ["Product Manager", "Negotiation", "Important ★★", 2, 12],
    ["Product Manager", "Problem Solving", "Important ★★", 2, 14],
    ["Product Manager", "Teamwork", "Nice-to-have ★", 1, 12],
]

for row_num, req in enumerate(pos_reqs, 2):
    for col, val in enumerate(req, 1):
        cell = ws3.cell(row=row_num, column=col, value=val)
        cell.border = thin_border if val != "" else None

for col in range(1, 6):
    ws3.column_dimensions[get_column_letter(col)].width = [22, 18, 18, 14, 10][col-1]

# ===== SHEET 4: Overall Score Calculator =====
ws4 = wb.create_sheet("Overall_Score")

# Header
ws4.cell(row=1, column=1, value="OVERALL SCORE CALCULATOR").font = Font(bold=True, size=14)
ws4.merge_cells('A1:E1')

# Formula explanation
ws4.cell(row=3, column=1, value="Formula: Σ(value × weight) / Σ(weight) × 5")
ws4.cell(row=4, column=1, value="Enter player attribute values in column C")

# Table headers
overall_headers = ["Attribute", "Weight", "Value (1-20)", "Multiplier", "Weighted Value"]
for col, header in enumerate(overall_headers, 1):
    cell = ws4.cell(row=6, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

# Data rows with formulas
overall_data = [
    ["Leadership", "High", 14, "=IF(B7=\"High\",1.5,IF(B7=\"Medium\",1.25,1))"],
    ["Teamwork", "High", 16, "=IF(B8=\"High\",1.5,IF(B8=\"Medium\",1.25,1))"],
    ["Communication", "Medium", 13, "=IF(B9=\"High\",1.5,IF(B9=\"Medium\",1.25,1))"],
    ["Coding", "High", 17, "=IF(B10=\"High\",1.5,IF(B10=\"Medium\",1.25,1))"],
    ["System Design", "High", 15, "=IF(B11=\"High\",1.5,IF(B11=\"Medium\",1.25,1))"],
    ["Problem Solving", "Medium", 15, "=IF(B12=\"High\",1.5,IF(B12=\"Medium\",1.25,1))"],
    ["Adaptability", "Medium", 12, "=IF(B13=\"High\",1.5,IF(B13=\"Medium\",1.25,1))"],
    ["Creativity", "Low", 11, "=IF(B14=\"High\",1.5,IF(B14=\"Medium\",1.25,1))"],
]

for row_num, data in enumerate(overall_data, 7):
    ws4.cell(row=row_num, column=1, value=data[0]).border = thin_border
    ws4.cell(row=row_num, column=2, value=data[1]).border = thin_border
    ws4.cell(row=row_num, column=3, value=data[2]).border = thin_border
    ws4.cell(row=row_num, column=4, value=data[3]).border = thin_border
    ws4.cell(row=row_num, column=5, value=f"=C{row_num}*D{row_num}").border = thin_border

# Totals and result
ws4.cell(row=16, column=3, value="Totals:").font = subheader_font
ws4.cell(row=16, column=4, value="=SUM(D7:D14)")
ws4.cell(row=16, column=5, value="=SUM(E7:E14)")

ws4.cell(row=18, column=1, value="OVERALL SCORE:").font = Font(bold=True, size=12)
ws4.cell(row=18, column=3, value="=ROUND(E16/D16*5,0)")
ws4.cell(row=18, column=3).font = Font(bold=True, size=16, color="2563EB")

ws4.cell(row=20, column=1, value="Tier:")
ws4.cell(row=20, column=3, value='=IF(C18>=90,"🥇 Gold",IF(C18>=80,"🥈 Silver","🥉 Bronze"))')

for col in range(1, 6):
    ws4.column_dimensions[get_column_letter(col)].width = [18, 10, 14, 12, 16][col-1]

# ===== SHEET 5: Fit Score Calculator =====
ws5 = wb.create_sheet("Fit_Score")

# Header
ws5.cell(row=1, column=1, value="POSITION FIT SCORE CALCULATOR").font = Font(bold=True, size=14)
ws5.merge_cells('A1:G1')

ws5.cell(row=3, column=1, value="Position: Senior Backend Developer")
ws5.cell(row=4, column=1, value="Formula: Σ(score × priority_weight) / Σ(priority_weight)")

# Table headers
fit_headers = ["Attribute", "Min Req", "Actual", "Gap", "Classification", "Score", "Priority Wgt", "Weighted Score"]
for col, header in enumerate(fit_headers, 1):
    cell = ws5.cell(row=6, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

# Data rows with formulas
fit_data = [
    ["Coding", 15, 17, "=C7-B7", '=IF(C7>=B7+2,"Exceeds",IF(C7>=B7-1,"Meets","Below"))', '=IF(C7>=B7+2,100,IF(C7>=B7-1,100,MAX(0,100-(B7-C7)*15)))', 3],
    ["System Design", 13, 15, "=C8-B8", '=IF(C8>=B8+2,"Exceeds",IF(C8>=B8-1,"Meets","Below"))', '=IF(C8>=B8+2,100,IF(C8>=B8-1,100,MAX(0,100-(B8-C8)*15)))', 3],
    ["Teamwork", 14, 14, "=C9-B9", '=IF(C9>=B9+2,"Exceeds",IF(C9>=B9-1,"Meets","Below"))', '=IF(C9>=B9+2,100,IF(C9>=B9-1,100,MAX(0,100-(B9-C9)*15)))', 3],
    ["Leadership", 12, 14, "=C10-B10", '=IF(C10>=B10+2,"Exceeds",IF(C10>=B10-1,"Meets","Below"))', '=IF(C10>=B10+2,100,IF(C10>=B10-1,100,MAX(0,100-(B10-C10)*15)))', 2],
    ["Problem Solving", 12, 10, "=C11-B11", '=IF(C11>=B11+2,"Exceeds",IF(C11>=B11-1,"Meets","Below"))', '=IF(C11>=B11+2,100,IF(C11>=B11-1,100,MAX(0,100-(B11-C11)*15)))', 2],
    ["Communication", 10, 13, "=C12-B12", '=IF(C12>=B12+2,"Exceeds",IF(C12>=B12-1,"Meets","Below"))', '=IF(C12>=B12+2,100,IF(C12>=B12-1,100,MAX(0,100-(B12-C12)*15)))', 1],
]

for row_num, data in enumerate(fit_data, 7):
    for col, val in enumerate(data, 1):
        cell = ws5.cell(row=row_num, column=col, value=val)
        cell.border = thin_border
    # Weighted score formula
    ws5.cell(row=row_num, column=8, value=f"=F{row_num}*G{row_num}").border = thin_border

# Totals and result
ws5.cell(row=14, column=6, value="Totals:")
ws5.cell(row=14, column=7, value="=SUM(G7:G12)")
ws5.cell(row=14, column=8, value="=SUM(H7:H12)")

ws5.cell(row=16, column=1, value="FIT SCORE:").font = Font(bold=True, size=12)
ws5.cell(row=16, column=3, value="=ROUND(H14/G14,0)&\"%\"")
ws5.cell(row=16, column=3).font = Font(bold=True, size=16, color="22C55E")

ws5.cell(row=18, column=1, value="Rating:")
ws5.cell(row=18, column=3, value='=IF(H14/G14>=90,"Excellent",IF(H14/G14>=70,"Good",IF(H14/G14>=50,"Fair","Poor")))')

# Gap analysis
ws5.cell(row=20, column=1, value="GAP ANALYSIS:").font = subheader_font
ws5.cell(row=21, column=1, value="▲ Exceeds:")
ws5.cell(row=21, column=2, value='=COUNTIF(E7:E12,"Exceeds")')
ws5.cell(row=22, column=1, value="— Meets:")
ws5.cell(row=22, column=2, value='=COUNTIF(E7:E12,"Meets")')
ws5.cell(row=23, column=1, value="▼ Below:")
ws5.cell(row=23, column=2, value='=COUNTIF(E7:E12,"Below")')

for col in range(1, 9):
    ws5.column_dimensions[get_column_letter(col)].width = [16, 8, 8, 6, 14, 8, 12, 14][col-1]

# ===== SHEET 6: Example =====
ws6 = wb.create_sheet("Example")

ws6.cell(row=1, column=1, value="COMPLETE EXAMPLE: Alex Chen - Tech Lead").font = Font(bold=True, size=14)

example_content = [
    [""],
    ["PLAYER ATTRIBUTES:"],
    ["Attribute", "Value", "Library Weight", "Multiplier"],
    ["Leadership", 17, "High", 1.5],
    ["Teamwork", 14, "High", 1.5],
    ["Communication", 16, "Medium", 1.25],
    ["Coding", 18, "High", 1.5],
    ["System Design", 17, "High", 1.5],
    ["Problem Solving", 17, "Medium", 1.25],
    ["Adaptability", 13, "Medium", 1.25],
    ["Initiative", 16, "Low", 1.0],
    [""],
    ["OVERALL SCORE CALCULATION:"],
    ["Sum of weighted values:", "= 17×1.5 + 14×1.5 + 16×1.25 + 18×1.5 + 17×1.5 + 17×1.25 + 13×1.25 + 16×1.0 = 173.75"],
    ["Sum of weights:", "= 1.5 + 1.5 + 1.25 + 1.5 + 1.5 + 1.25 + 1.25 + 1.0 = 10.75"],
    ["Raw average:", "= 173.75 / 10.75 = 16.16"],
    ["Overall Score:", "= 16.16 × 5 = 80.8 ≈ 81 (Silver Tier)"],
    [""],
    ["POSITION FIT (Tech Lead):"],
    ["Tech Lead requires: Leadership(14), Teamwork(15), Communication(12), Coding(16), Problem(14)"],
    [""],
    ["Attribute", "Actual", "Required", "Gap", "Score"],
    ["Leadership", 17, 14, "+3 Exceeds", 100],
    ["Teamwork", 14, 15, "-1 Meets", 100],
    ["Communication", 16, 12, "+4 Exceeds", 100],
    ["Coding", 18, 16, "+2 Exceeds", 100],
    ["Problem Solving", 17, 14, "+3 Exceeds", 100],
    [""],
    ["Fit Score: (100×3 + 100×3 + 100×2 + 100×3 + 100×2) / (3+3+2+3+2) = 1300/13 = 100%"],
]

for row_num, row_data in enumerate(example_content, 1):
    if isinstance(row_data, list):
        for col, val in enumerate(row_data, 1):
            ws6.cell(row=row_num, column=col, value=val)
    else:
        ws6.cell(row=row_num, column=1, value=row_data)

ws6.column_dimensions['A'].width = 20
ws6.column_dimensions['B'].width = 60

# Save the workbook
output_path = "/Users/gdrom/Desktop/allkons/ascend-hr-docs/ascendhr/engine/player-card-calculations.xlsx"
wb.save(output_path)
print(f"Excel file created: {output_path}")
